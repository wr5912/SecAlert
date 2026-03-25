"""Excel 报表导出器

使用 openpyxl 生成 xlsx 格式报表
"""

from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


def generate_excel_report(
    metrics: Dict[str, Any],
    report_type: str = "daily"
) -> bytes:
    """生成 Excel 报表

    Args:
        metrics: 报表数据
        report_type: "daily" 或 "weekly"

    Returns:
        xlsx 字节数据
    """
    wb = Workbook()

    # Sheet 1: 摘要
    ws_summary = wb.active
    ws_summary.title = "摘要"

    # 样式定义
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    headers = ["指标", "数值"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 写入数据行
    summary_data = [
        ("日期", metrics.get("date", "")),
        ("总告警", metrics.get("total_alerts", 0)),
        ("总链数", metrics.get("total_chains", 0)),
        ("真威胁", metrics.get("true_positives", 0)),
        ("误报", metrics.get("false_positives", 0)),
        ("误报率", f"{metrics.get('fp_rate', 0) * 100:.1f}%"),
    ]

    for row, (key, value) in enumerate(summary_data, 2):
        cell1 = ws_summary.cell(row=row, column=1, value=key)
        cell1.border = thin_border
        cell1.alignment = Alignment(horizontal='left', vertical='center')
        cell2 = ws_summary.cell(row=row, column=2, value=value)
        cell2.border = thin_border
        cell2.alignment = Alignment(horizontal='right', vertical='center')

    # 自动列宽
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 15

    # Sheet 2: 每日明细 (用于周报)
    if report_type == "weekly" and "daily_breakdown" in metrics:
        ws_detail = wb.create_sheet("每日明细")

        # 写入明细表头
        detail_headers = ["日期", "总告警", "总链数", "真威胁", "误报", "误报率"]
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入每日数据
        for row_idx, day_data in enumerate(metrics["daily_breakdown"], 2):
            ws_detail.cell(row=row_idx, column=1, value=day_data.get("date", "")).border = thin_border
            ws_detail.cell(row=row_idx, column=2, value=day_data.get("total_alerts", 0)).border = thin_border
            ws_detail.cell(row=row_idx, column=3, value=day_data.get("total_chains", 0)).border = thin_border
            ws_detail.cell(row=row_idx, column=4, value=day_data.get("true_positives", 0)).border = thin_border
            ws_detail.cell(row=row_idx, column=5, value=day_data.get("false_positives", 0)).border = thin_border
            ws_detail.cell(row=row_idx, column=6, value=f"{day_data.get('fp_rate', 0) * 100:.1f}%").border = thin_border

        # 自动列宽
        for col in range(1, 7):
            ws_detail.column_dimensions[chr(64 + col)].width = 12

    # Sheet 3: 严重度分布
    ws_severity = wb.create_sheet("严重度分布")
    ws_severity.cell(row=1, column=1, value="严重度").fill = header_fill
    ws_severity.cell(row=1, column=1, value="严重度").font = header_font
    ws_severity.cell(row=1, column=1).border = thin_border
    ws_severity.cell(row=1, column=2, value="数量").fill = header_fill
    ws_severity.cell(row=1, column=2, value="数量").font = header_font
    ws_severity.cell(row=1, column=2).border = thin_border

    severity_dist = metrics.get("severity_distribution", {})
    severity_labels = {"critical": "严重", "high": "高", "medium": "中", "low": "低"}
    for row_idx, (key, label) in enumerate(severity_labels.items(), 2):
        ws_severity.cell(row=row_idx, column=1, value=label).border = thin_border
        ws_severity.cell(row=row_idx, column=2, value=severity_dist.get(key, 0)).border = thin_border

    ws_severity.column_dimensions['A'].width = 12
    ws_severity.column_dimensions['B'].width = 12

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
